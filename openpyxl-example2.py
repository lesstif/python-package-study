import math
import constants

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = constants.EXCEL_TEST_WOOKBOOK

## 1번째 워크시트 생성
ws1 = wb.active
ws1.title = constants.EXCEL_TEST_WB_TITLE1

nations = {
    '대한민국': '서울',
    '미국': '워싱턴',
    '일본': '도쿄',
    '아프가니스탄': '카불',
}

heading = [ '국가', '수도' ]
ws1.append(heading)

for n in nations.items():
    ws1.append(n)

## 2번째 워크시트 생성
ws2 = wb.create_sheet(title="Pi")

ws2['F1'] = '파이값'
ws2['F5'] = math.pi

## 3번째 워크시트 생성
ws3 = wb.create_sheet(title='Data')
for row in range(10, 20):
    for col in range(27, 54):
        cl = get_column_letter(col)
        _ = ws3.cell(column=col, row=row, value=f'{cl}{row}')

print(ws3['AA10'].value)

wb.save(filename=dest_filename)